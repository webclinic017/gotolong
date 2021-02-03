# Create your views here.

# To address, NameError: name 'Sum' is not defined
from django.db.models import (Sum, Count)

from django.db.models.functions import (ExtractYear, Round, ExtractMonth)

from django.views.generic.list import ListView
from django.views.generic.dates import YearArchiveView, MonthArchiveView

from django_gotolong.bstmtdiv.models import BstmtDiv

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update

import pandas as pd
import csv, io
# works only with .xlsx
import openpyxl
import os
import calendar

# for xls (openpyxl doesn't work with xls)
# import pyexcel_xls

from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseRedirect


class BstmtDivYearArchiveView(YearArchiveView):
    queryset = BstmtDiv.objects.all()
    date_field = "bsdiv_date"
    make_object_list = True
    allow_future = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        total_amount = round(
            BstmtDiv.objects.all().filter(bsdiv_date__year=self.get_year()).aggregate(Sum('bsdiv_amount'))[
                'bsdiv_amount__sum'])
        summary_list = (
            BstmtDiv.objects.all().filter(bsdiv_date__year=self.get_year()).annotate(
                month=ExtractMonth('bsdiv_date')).values('month').annotate(
                Total=Round(Sum('bsdiv_amount'))))
        context["total_amount"] = total_amount
        context["summary_list"] = summary_list
        return context


class BstmtDivMonthArchiveView(MonthArchiveView):
    queryset = BstmtDiv.objects.all()
    date_field = "bsdiv_date"
    make_object_list = True
    allow_future = True


class BstmtDivListView(ListView):
    model = BstmtDiv
    # if pagination is desired
    # paginate_by = 300
    queryset = BstmtDiv.objects.all()

    year_list = BstmtDiv.objects.dates('bsdiv_date', 'year')
    month_list = BstmtDiv.objects.dates('bsdiv_date', 'month')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["year_list"] = self.year_list
        context["month_list"] = self.month_list
        # aggregate returns dictionary
        # get value from dictionary
        total_amount = round(BstmtDiv.objects.all().aggregate(Sum('bsdiv_amount'))[
                                 'bsdiv_amount__sum'])
        summary_list = (
            BstmtDiv.objects.all().annotate(
                year=ExtractYear('bsdiv_date')).values('year').annotate(
                Total=Round(Sum('bsdiv_amount'))))
        month_summary_list = (
            BstmtDiv.objects.all().annotate(
                month=ExtractMonth('bsdiv_date')).values('month').annotate(
                Total=Round(Sum('bsdiv_amount')))).order_by('month')
        context["total_amount"] = total_amount
        context["summary_list"] = summary_list
        context["month_summary_list"] = month_summary_list
        return context


class BstmtDivAmountView(ListView):
    model = BstmtDiv
    # if pagination is desired
    # paginate_by = 300
    queryset = BstmtDiv.objects.all()

    year_list = BstmtDiv.objects.dates('bsdiv_date', 'year')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["year_list"] = self.year_list
        # aggregate returns dictionary
        # get value from dictionary
        summary_list = (
            BstmtDiv.objects.all().values('ticker')
                .annotate(Total=Round(Sum('bsdiv_amount')))).order_by('-Total')
        context["summary_list"] = summary_list
        return context


class BstmtDivFrequencyView(ListView):
    model = BstmtDiv
    # if pagination is desired
    # paginate_by = 300
    queryset = BstmtDiv.objects.all()

    year_list = BstmtDiv.objects.dates('bsdiv_date', 'year')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["year_list"] = self.year_list
        # aggregate returns dictionary
        # get value from dictionary
        summary_list = (
            BstmtDiv.objects.all().values('bsdiv_ticker')
                .annotate(Total=Round(Count('bsdiv_ticker')))).order_by('-Total')
        context["summary_list"] = summary_list
        return context


def div_date_iso(self, txn_date):
    month_name_abbr_to_num_dict = {name: num for num, name in enumerate(calendar.month_abbr) if num}

    try:
        # dd-mmm-yy
        txn_date_arr = txn_date.split('/')
        txn_day = txn_date_arr[0].strip()
        txn_month = txn_date_arr[1].strip()
        txn_year = txn_date_arr[2].strip()
        if txn_month.isdigit():
            # get rid of leading 0 in month number
            txn_month = str(int(txn_month))
        else:
            # month name to number
            txn_month = str(month_name_abbr_to_num_dict[txn_month])
        txn_date_iso = txn_year + "-" + txn_month + "-" + txn_day
        # ignore rest
    except ValueError:
        logging.error('ValueError ', txn_date)
    except IndexError:
        logging.error('IndexError ', txn_date)

    return txn_date_iso


# one parameter named request
def bstmtdiv_upload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "bstmtdiv/bstmtdiv_list.html"
    data = BstmtDiv.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    bank_name = request.POST.get('bank')
    print(bank_name)

    # delete existing records
    print('Deleted existing BstmtDiv data')
    BstmtDiv.objects.all().delete()

    # req_file = request.FILES['file']
    unique_id = 0
    for req_file in request.FILES.getlist("files"):

        # let's check if it is a csv file
        if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
            # get worksheet name
            # print('temporary file path:', req_file.temporary_file_path)
            print(req_file)

            try:
                if req_file.name.endswith('.xlsx'):
                    if True:
                        print('openpyxl')
                        wb = openpyxl.load_workbook(req_file)
                        print(wb.sheetnames)
                        sheet_name = wb.sheetnames[0]
                        print(sheet_name)
                        ws = wb[sheet_name]
                        df = pd.DataFrame(ws.values)
                    else:
                        print('pd.ExcelFile')
                        # not supported on InMemoryUploadedFile
                        # file_loc = os.path.dirname(os.path.realpath(req_file))
                        # print ('file location', file_loc)
                        xl = pd.ExcelFile(req_file.read())
                        if debug_level > 0:
                            print(xl.sheet_names)
                        # single worksheet - Data
                        sheet_name = xl.sheet_names[0]
                        df = xl.parse(sheet_name)

                if req_file.name.endswith('.xls'):
                    method = "panda"
                    print('xls parsed using', method)
                    if method == 'xlrd':
                        print('x')
                        # sheet  = pyexcel.get_sheet(req_file)
                        # df = sheet.content
                    elif method == "xlrd":
                        print('y')
                        workbook = xlrd.open_workbook(file_contents=req_file.read())
                        sheet = workbook.sheet_by_index(0)
                    elif method == "panda":
                        print('pd.read_excel')
                        # not supported on InMemoryUploadedFile
                        # file_loc = os.path.dirname(os.path.realpath(req_file))
                        # print ('file location', file_loc)
                        df = pd.read_excel(req_file.read())
            except Exception as e:
                print(e)

            # find index of 'Transactions List -' in Unnamed:1 column
            # top_index = map(lambda x: x.find('Transactions List -'), df['Unnamed: 1'])

            # print("top_index : " + str(top_index))
            # find index of 'Legends Used in Account Statement' in Unnamed:1 column
            # bottom_index = map(lambda x: x.find('Legends Used in Account Statement'), df['Unnamed: 1'])

            # print("bottom_index : " + str(bottom_index))

            # remove top 11 lines (ignore logo) from dataframe
            # df = df.iloc[top_index:]

            # remove 28 lines from bottom  : exclude Legends data
            # df = df[:bottom_index-1]

            if bank_name == 'icici':
                # remarks - description
                df.columns = (
                    'ignore_0', 'serial_num', 'value_date', 'txn_date', 'ref_cheque_num', 'txn_description',
                    'withdraw_amount',
                    'deposit_amount', 'balance')
            elif bank_name == 'sbi':
                # credit -> deposit
                # debit -> withdraw
                df.columns = (
                    'txn_date', 'value_date', 'txn_description', 'ref_cheque_num', 'withdraw_amount', 'deposit_amount',
                    'balance')
            elif bank_name == 'hdfc':
                # narration -> description
                df.columns = (
                    'txn_date', 'txn_description', 'ref_cheque_num', 'value_date', 'withdraw_amount', 'deposit_amount',
                    'balance')
            elif bank_name == 'axis':
                # SRL NO, Tran Date, CHQNO, PARTICULARS, DR, CR, BAL, SOL
                df.columns = (
                    'serial_num', 'txn_date', 'ref_cheque_num', 'txn_description', 'withdraw_amount', 'deposit_amount',
                    'balance', 'sol')
            else:
                print('unsupported bank: ', bank_name)
                print('please configure df.columns')

            # TODO : Remove lines in case last column is null. Use -1 instead
            # if bank_name == 'icici-bank':
            # Keep only if 'Balance (INR)' column is not NA
            #    df = df[df['Unnamed: 8'].notnull()]
            # elif bank_name == 'sbi-bank':
            #    # Keep only if 'Balance' column is not NA
            #    df = df[df['Unnamed: 6'].notnull()]
            # elif bank_name == 'hdfc-bank':
            #    # Keep only if 'Closing Balance' column is not NA
            #    df = df[df['Unnamed: 6'].notnull()]

            # keep only if balance is non-null
            df = df[df['balance'].notnull()]

            # txn_description is non-null
            df = df[df['txn_description'].notnull()]

            # select only the required columns : txn date, description and deposit amount
            df = df[['txn_date', 'txn_description', 'deposit_amount']]

            # invalid data processing
            if bank_name == 'icici-bank':
                # first line is 'Transaction Date from' : case sensitive
                # df = df[df['Unnamed: 1'] != 'Transaction Date from']
                print('not used now : Transaction Date from')

            # first column processing
            if bank_name == 'icici-bank':
                # remove first column (NaN before S.No) - axis 1 is column, 0 is row
                # df.drop(df.columns[0], axis=1, inplace=True)
                print('no longer need to remove first column explicitly')

            if bank_name == 'icici-bank' or bank_name == 'sbi-bank':
                # remove last column : Balance (INR) - ICICI
                # remove last column : Balance - SBI
                # df = df.iloc[:, :-1]
                print('no longer need to remove last column explicitly')

            # change columns from second line from top
            # df.columns = df.iloc[0]

            if debug_level > 0:
                print("columns : " + df.columns)

            # remove the top line that contains column name
            df = df.iloc[1:]

            # For backward compatibility
            # Currently, clear the value in last column  : Balance(INR)
            # df['Balance (INR)'] = 0

            # Add a dummy column - for backward compatibility
            # df[''] = ''

            # Keep only ACH and CMS for dividend transfer
            # remove transactions with remarks ':Int.Pd:' and 'BY CASH', MMT (IMPS), EBA etc
            # Donot use single quote around ppatern for regex

            # 2015 and before - credit using ECS
            # 2016 onwards - ACH/CMS etc
            pattern = "ACH/|CMS/|ECS "
            filter = df['txn_description'].str.contains(pattern, regex=True)

            if debug_level > 0:
                print(filter)

            df = df[filter]

            # fix formatting for date : dd-mm-yy to dd/mm/yy
            # for any bank like SBI bank
            df['txn_date'] = df['txn_date'].replace({'-': '/'})

            # TODO: how to convert mmm to mm ?
            # TODO: how to convert yy to yyyy

            # if bank_name == 'icici-bank':
            # dd/mm/YYYY
            # print('date already in format')
            # elif bank_name == 'hdfc-bank':
            # convert dd/mm/yy to dd/mm/yyyy
            # d = datetime.datetime.strptime(my_date, '%d/%m/%y')
            # d.strftime('%d/%m/%Y')
            # df['txn_date'] = df['txn_date'].dt.strftime('%d/%m/%Y')
            # print('try to convert yy to yyyy')
            # elif bank_name == 'sbi-bank':
            # convert dd-mmm-yy to dd/mm/yyyy
            # d = datetime.datetime.strptime(my_date, '%d-%b-%y')
            # d.strftime('%d/%m/%Y')
            # df['txn_date'] = df['txn_date'].dt.strftime('%d/%m/%Y')
            # print('convert dd-mmm-yy to dd/mm/yyyy')

            if debug_level > 0:
                print(df)

            data_set = df.to_csv(header=True, index=False)

            # df.to_csv(csv_file, header=False, index=False)

            # df = pd.read_excel(excel_file, sheetName=None)
            # print df.keys()

            # pd.read_excel(excel_file).to_csv(csv_file, index=False)

        if req_file.name.endswith('.csv'):
            data_set = req_file.read().decode('UTF-8')

        if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
            messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
            return HttpResponseRedirect(reverse("bstmtdiv-list"))

        # setup a stream which is when we loop through each line we are able to handle a data in a stream

        io_string = io.StringIO(data_set)
        next(io_string)

        for column in csv.reader(io_string, delimiter=',', quotechar='"'):
            unique_id += 1
            bsdiv_date = div_date_iso(request, column[0].strip())
            bsdiv_remarks = column[1].strip()
            bsdiv_amount = column[2].strip()

            # convert dd-mmm-yy to YYYY-mm-dd
            # txn_date = txn_date_iso(request, column[12])

            print(unique_id, column)

            _, created = BstmtDiv.objects.update_or_create(
                bsdiv_id=unique_id,
                bsdiv_date=bsdiv_date,
                bsdiv_remarks=bsdiv_remarks,
                bsdiv_amount=bsdiv_amount
            )
    # context = {}
    # render(request, template, context)
    #
    lastrefd_update("bstmtdiv")
    #
    print('Completed loading new BstmtDiv data')
    return HttpResponseRedirect(reverse("bstmtdiv-list"))

# from django.http import HttpResponse
# def index(request):
#    return HttpResponse("Hello, world. You're at the polls index.")
#
