from django.shortcuts import render

# Create your views here.

from django.views.generic.list import ListView

# from django_filters.rest_framework import DjangoFilterBackend, FilterSet, OrderingFilter

import pandas as pd
import csv, io
import openpyxl

from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseRedirect

from django.db.models import OuterRef, Subquery
from django.db.models import IntegerField, FloatField, ExpressionWrapper
from django.db.models import F, Count, Sum, Max, Min

from django_gotolong.amfi.models import Amfi

from django_gotolong.dematsum.models import DematSum
from django_gotolong.gweight.models import Gweight

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update


class AmfiListView(ListView):
    model = Amfi
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Amfi.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class AmfiAmountView(ListView):
    dematsum_qs = DematSum.objects.filter(ds_isin=OuterRef("comp_isin"))
    gweight_qs = Gweight.objects.filter(gw_cap_type=OuterRef("cap_type"))
    queryset = Amfi.objects.all(). \
        annotate(value_cost=Subquery(dematsum_qs.values('ds_costvalue')[:1])). \
        annotate(cap_weight=Subquery(gweight_qs.values('gw_cap_weight')[:1])). \
        annotate(deficit=ExpressionWrapper(F('cap_weight') * 1000 - F('value_cost'), output_field=IntegerField())). \
        values('comp_rank', 'comp_name', 'value_cost', 'deficit'). \
        order_by('comp_rank')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context


class AmfiPortfWeightView(ListView):

    def get_queryset(self):
        sum_portf_insts = DematSum.objects.filter(ds_user_id=self.request.user.id). \
            aggregate(Sum('ds_mktvalue'))
        print('first:', sum_portf_insts)
        try:
            print('second', sum_portf_insts['ds_mktvalue__sum'])
            sum_portf = round(sum_portf_insts['ds_mktvalue__sum'])
            print('DS: sum portf ', sum_portf)
        except KeyError:
            sum_portf = 0
            print('DS: sum portf ', sum_portf)

        dematsum_qs = DematSum.objects.filter(ds_user_id=self.request.user.id). \
            filter(ds_isin=OuterRef("comp_isin"))
        queryset = Amfi.objects.all(). \
            annotate(mkt_value=Subquery(dematsum_qs.values('ds_mktvalue')[:1])). \
            annotate(portf_weight=ExpressionWrapper(F('mkt_value') * 100.0 / sum_portf, output_field=FloatField())). \
            exclude(portf_weight__isnull=True). \
            exclude(portf_weight__lt=0). \
            values('comp_rank', 'comp_name', 'mkt_value', 'portf_weight'). \
            order_by('comp_rank')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

class AmfiDeficitView(ListView):
    dematsum_qs = DematSum.objects.filter(ds_isin=OuterRef("comp_isin"))
    gweight_qs = Gweight.objects.filter(gw_cap_type=OuterRef("cap_type"))
    queryset = Amfi.objects.all(). \
        annotate(value_cost=Subquery(dematsum_qs.values('ds_costvalue')[:1])). \
        annotate(cap_weight=Subquery(gweight_qs.values('gw_cap_weight')[:1])). \
        annotate(deficit=ExpressionWrapper(F('cap_weight') * 1000 - F('value_cost'), output_field=IntegerField())). \
        values('comp_rank', 'comp_name', 'value_cost', 'deficit'). \
        filter(value_cost__isnull=False). \
        order_by('-deficit')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context


class AmfiNotableExclusionView(ListView):
    dematsum_qs = DematSum.objects.filter(ds_isin=OuterRef("comp_isin"))
    gweight_qs = Gweight.objects.filter(gw_cap_type=OuterRef("cap_type"))
    # missing large cap and mid cap : top 250 only
    queryset = Amfi.objects.all(). \
        annotate(value_cost=Subquery(dematsum_qs.values('ds_costvalue')[:1])). \
        annotate(cap_weight=Subquery(gweight_qs.values('gw_cap_weight')[:1])). \
        annotate(deficit=ExpressionWrapper(F('cap_weight') * 1000 - F('value_cost'), output_field=IntegerField())). \
        values('comp_rank', 'comp_name', 'value_cost', 'deficit'). \
        filter(value_cost__isnull=True).filter(comp_rank__lte=250).order_by('comp_rank')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context


class AmfiNotableInclusionView(ListView):
    dematsum_qs = DematSum.objects.filter(ds_isin=OuterRef("comp_isin"))
    gweight_qs = Gweight.objects.filter(gw_cap_type=OuterRef("cap_type"))
    # included nano cap and micro cop > 500 only
    queryset = Amfi.objects.all(). \
        annotate(value_cost=Subquery(dematsum_qs.values('ds_costvalue')[:1])). \
        annotate(cap_weight=Subquery(gweight_qs.values('gw_cap_weight')[:1])). \
        annotate(deficit=ExpressionWrapper(F('cap_weight') * 1000 - F('value_cost'), output_field=IntegerField())). \
        values('comp_rank', 'comp_name', 'value_cost', 'deficit'). \
        filter(value_cost__isnull=False).filter(comp_rank__gte=500).order_by('-comp_rank')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        return context

# one parameter named request
def amfi_upload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "amfi/amfi_list.html"
    data = Amfi.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        if sheet_name != 'Data':
            print("sheet name changed to", sheet_name)

        # ignore top two line : Average Market Capitalization of listed companies during the six months ended 
        # remove top two line from dataframe
        df = df.iloc[2:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # change column name of data frame 
        columns_list = ['sr_no', 'name', 'isin', 'bse_symbol', 'bse_mcap',
                        'nse_symbol', 'nse_mcap', 'mse_symbol', 'mse_mcap',
                        'avg_mcap', 'captype']
        df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        df[["avg_mcap"]] = df[["avg_mcap"]].astype(int)

        # drop columns that are not required
        skip_columns_list = ['bse_mcap', 'nse_mcap', 'mse_symbol', 'mse_mcap']
        df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
        return HttpResponseRedirect(reverse("amfi-list"))

    # delete existing records
    print('Deleted existing Amfi data')
    Amfi.objects.all().delete()

    # setup a stream which is when we loop through each line we are able to handle a data in a stream

    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar='"'):
        column[0] = column[0].strip()
        column[1] = column[1].strip()

        comp_rank = int(column[0])
        cap_type = column[6]
        if comp_rank > 500 and cap_type == 'Small Cap':
            if comp_rank > 750:
                cap_type = 'Nano Cap'
            else:
                cap_type = 'Micro Cap'

        _, created = Amfi.objects.update_or_create(
            comp_rank=comp_rank,
            comp_name=column[1],
            comp_isin=column[2],
            bse_symbol=column[3],
            nse_symbol=column[4],
            avg_mcap=column[5],
            cap_type=cap_type
        )

    lastrefd_update("amfi")

    print('Completed loading new Amfi data')
    return HttpResponseRedirect(reverse("amfi-list"))

# from django.http import HttpResponse
# def index(request):
#    return HttpResponse("Hello, world. You're at the polls index.")
#
