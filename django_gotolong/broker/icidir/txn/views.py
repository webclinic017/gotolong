# Create your views here.

from django_gotolong.broker.icidir.txn.models import BrokerIcidirTxn

from django.utils import timezone

from django.views.generic.list import ListView
from django.views.generic.dates import YearArchiveView, MonthArchiveView

from django.db.models import IntegerField, F, ExpressionWrapper, fields, Max, Min, Sum, Count
from django.db.models.functions import (ExtractYear, Round, ExtractMonth)
from django.db.models.expressions import RawSQL

import calendar
import pandas as pd
import csv, io
import openpyxl
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseRedirect

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update


class BrokerIcidirTxnYearArchiveView(YearArchiveView):
    queryset = BrokerIcidirTxn.objects.all()
    date_field = "txn_date"
    make_object_list = True
    allow_future = True

    year_list = BrokerIcidirTxn.objects.dates('txn_date', 'year')
    month_list = BrokerIcidirTxn.objects.dates('txn_date', 'month')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["year_list"] = self.year_list
        context["month_list"] = self.month_list
        txn_amount = ExpressionWrapper(F('quantity') * F('txn_price'), output_field=IntegerField())
        total_amount = round(BrokerIcidirTxn.objects.all().filter(txn_date__year=self.get_year()).
                             aggregate(txn_amount=Sum(txn_amount))['txn_amount'])

        summary_list = (
            BrokerIcidirTxn.objects.all().filter(txn_date__year=self.get_year()).
                annotate(
                month=ExtractMonth('txn_date')).values('month').annotate(
                Total=Round(Sum(txn_amount)))).order_by('month')
        context["total_amount"] = total_amount
        context["summary_list"] = summary_list
        return context


class BrokerIcidirTxnMonthArchiveView(MonthArchiveView):
    queryset = BrokerIcidirTxn.objects.all()
    date_field = "txn_date"
    make_object_list = True
    allow_future = True


class BrokerIcidirTxnListView(ListView):
    model = BrokerIcidirTxn

    # if pagination is desired
    # paginate_by = 300
    queryset = BrokerIcidirTxn.objects.all()

    year_list = BrokerIcidirTxn.objects.dates('txn_date', 'year')

    # month_list = BrokerIcidirTxn.objects.dates('txn_date', 'month')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["year_list"] = self.year_list
        #  context["month_list"] = self.month_list
        txn_amount = ExpressionWrapper(F('quantity') * F('txn_price'), output_field=IntegerField())
        total_amount = round(BrokerIcidirTxn.objects.all().aggregate(txn_amount=Sum(txn_amount))[
                                 'txn_amount'])
        summary_list = (
            BrokerIcidirTxn.objects.all().annotate(
                year=ExtractYear('txn_date')).values('year').annotate(
                Total=Round(Sum(txn_amount)))).order_by('year')
        context["total_amount"] = total_amount
        context["summary_list"] = summary_list
        return context


class BrokerIcidirTxnGapView(ListView):
    model = BrokerIcidirTxn

    # if pagination is desired
    # paginate_by = 300
    date_now = timezone.now()
    # date_now = datetime.today().strftime('%Y-%m-%d')
    time_diff = ExpressionWrapper(- F('txn_date'), output_field=fields.DurationField())
    # time_diff = ExpressionWrapper(F('due_date'')-Now())
    # ,min_txn_date = Min('txn_date')
    # days=ExpressionWrapper(date_now - F('txn_date'), output_field=fields.DurationField())
    # date_fmt = '%Y-%m-%d'
    # annotate(txn_date_fmt=RawSQL('DATE_FORMAT(txn_date, "%Y-%m-%d")',())).\
    # annotate(str_datetime=Cast('txn_date', CharField())). \
    #         annotate(str_datetime=Cast('txn_date', CharField())). \

    queryset = BrokerIcidirTxn.objects. \
        values('stock_symbol'). \
        annotate(max_txn_date=Max('txn_date')). \
        order_by('max_txn_date'). \
        annotate(months_gap=Min(RawSQL('TIMESTAMPDIFF(month, txn_date, curdate())', ()))). \
        order_by('-months_gap')

    # annotate(first_time_diff=RawSQL('TIMESTAMPDIFF(month, txn_date, curdate())', ())). \
    # annotate(time_diff=Max('first_time_diff')).\
    # annotate(max_txn_date=Cast('txn_date', CharField())).\
    # order_by('time_diff')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class BrokerIcidirTxnStatView(ListView):
    model = BrokerIcidirTxn

    # if pagination is desired
    # paginate_by = 300
    txn_amount = ExpressionWrapper(F('quantity') * F('txn_price'), output_field=IntegerField())
    queryset = BrokerIcidirTxn.objects.all(). \
        annotate(txn_year=ExtractYear('txn_date')). \
        values('txn_year').annotate(txn_amount=Round(Sum(txn_amount))). \
        order_by('txn_year')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


class BrokerIcidirTxnStatBuySellView(ListView):
    model = BrokerIcidirTxn

    # if pagination is desired
    # paginate_by = 300
    txn_amount = ExpressionWrapper(F('quantity') * F('txn_price'), output_field=IntegerField())
    queryset = BrokerIcidirTxn.objects.all(). \
        annotate(txn_year=ExtractYear('txn_date')). \
        values('txn_year', 'action').annotate(txn_amount=Round(Sum(txn_amount))). \
        order_by('txn_year')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


def txn_date_iso(self, txn_date):
    month_name_abbr_to_num_dict = {name: num for num, name in enumerate(calendar.month_abbr) if num}

    try:
        # dd-mmm-yy
        txn_date_arr = txn_date.split('-')
        txn_day = txn_date_arr[0].strip()
        txn_month = txn_date_arr[1].strip()
        txn_year = txn_date_arr[2].strip()
        if txn_month.isdigit():
            # get rid of leading 0 in month number
            txn_month = str(int(txn_month))
        else:
            # month name to number
            txn_month = str(month_name_abbr_to_num_dict[txn_month])
        txn_date_iso = txn_year + "-" + txn_month + "-" + txn_day
        # ignore rest
    except ValueError:
        logging.error('ValueError ', txn_date, row_list)
    except IndexError:
        logging.error('IndexError ', txn_date, row_list)

    return txn_date_iso


# one parameter named request
def BrokerIcidirTxnUpload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "broker/icidir/txn/demattxn_list.html"
    data = BrokerIcidirTxn.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        if sheet_name != 'Data':
            print("sheet name changed to", sheet_name)

        # ignore top two line : Average Market Capitalization of listed companies during the six months ended
        # remove top two line from dataframe
        df = df.iloc[2:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # change column name of data frame
        columns_list = ['stock_symbol', 'comp_name', 'isin_code', 'action', 'quantity',
                        'txn_price', 'brokerage', 'txn_charges', 'stamp_duty',
                        'segment', 'stt', 'remarks', 'txn_date',
                        'exchange', 'unused1']
        df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        df[["avg_mcap"]] = df[["avg_mcap"]].astype(int)

        # drop columns that are not required
        # skip_columns_list = ['bse_mcap', 'nse_mcap', 'mse_symbol', 'mse_mcap']
        # df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
        return HttpResponseRedirect(reverse("broker-icidir-txn-list"))

    # delete existing records
    print('Deleted existing BrokerIcidirTxn data')
    BrokerIcidirTxn.objects.all().delete()

    # setup a stream which is when we loop through each line we are able to handle a data in a stream

    io_string = io.StringIO(data_set)
    next(io_string)
    unique_id = 0
    for column in csv.reader(io_string, delimiter=',', quotechar='"'):
        unique_id += 1
        column[0] = column[0].strip()
        column[1] = column[1].strip()

        # convert dd-mmm-yy to YYYY-mm-dd
        txn_date = txn_date_iso(request, column[12])

        print(unique_id, column)

        _, created = BrokerIcidirTxn.objects.update_or_create(
            dt_id=unique_id,
            stock_symbol=column[0],
            comp_name=column[1],
            isin_code=column[2],
            action=column[3],
            quantity=column[4],
            txn_price=column[5],
            brokerage=column[6],
            txn_charges=column[7],
            stamp_duty=column[8],
            segment=column[9],
            stt=column[10],
            remarks=column[11],
            txn_date=txn_date,
            exchange=column[13],
            unused1=column[14]
        )
    # context = {}
    # render(request, template, context)
    #
    lastrefd_update("broker-icidir-txn")
    #
    print('Completed loading new BrokerIcidirTxn data')
    return HttpResponseRedirect(reverse("broker-icidir-txn-list"))

# from django.http import HttpResponse
# def index(request):
#    return HttpResponse("Hello, world. You're at the polls index.")
#
