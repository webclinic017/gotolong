# Create your views here.

from .models import BrokerZerodhaSum

from django.shortcuts import render, redirect, get_object_or_404

from django.views.generic.list import ListView

from django.db.models import OuterRef, Subquery, Count, Sum
from django.db.models.functions import Trim, Lower, Round

from django_gotolong.amfi.models import Amfi
from django_gotolong.gfundareco.models import Gfundareco

import pandas as pd
import csv, io
import openpyxl
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseRedirect

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update


class BrokerZerodhaSumListView(ListView):
    model = BrokerZerodhaSum

    # if pagination is desired
    # paginate_by = 300

    queryset = BrokerZerodhaSum.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


# one parameter named request
def BrokerZerodhaSumUpload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    # only POST method is supported with upload
    # avoid direct access to this url
    template = "invalid-get-request.html"
    data = BrokerZerodhaSum.objects.all()

    # GET request returns the value of the data with the specified key.
    print("method : ", request.method, template)
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        if sheet_name != 'Data':
            print("sheet name changed to", sheet_name)

        # ignore top two line : Average Market Capitalization of listed companies during the six months ended
        # remove top two line from dataframe
        df = df.iloc[2:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # change column name of data frame
        columns_list = ['bzs_instrument', 'bzs_quantity', 'bzs_average_cost', 'bzs_ltp',
                        'bzs_cur_value', 'bzs_pnl', 'bzs_net_chg', 'bzs_day_chg']
        df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        df[["avg_mcap"]] = df[["avg_mcap"]].astype(int)

        # drop columns that are not required
        # skip_columns_list = ['bse_mcap', 'nse_mcap', 'mse_symbol', 'mse_mcap']
        # df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
        return HttpResponseRedirect(reverse("broker-zerodha-sum-list"))

    # delete existing records
    print('Deleted existing BrokerZerodhaSum data')
    BrokerZerodhaSum.objects.all().delete()

    # setup a stream which is when we loop through each line we are able to handle a data in a stream

    io_string = io.StringIO(data_set)
    next(io_string)
    for column in csv.reader(io_string, delimiter=',', quotechar='"'):
        column[0] = column[0].strip()
        column[1] = column[1].strip()

        _, created = BrokerZerodhaSum.objects.update_or_create(
            bzs_instrument=column[0],
            bzs_quantity=column[1],
            bzs_average_cost=column[2],
            bzs_ltp=column[3],
            bzs_cur_value=column[4],
            bzs_pnl=column[5],
            bzs_net_chg=column[6],
            bzs_day_chg=column[7]
        )
    # context = {}
    # render(request, template, context)
    lastrefd_update("broker-zerodha-sum")
    #
    print('Completed loading new BrokerZerodhaSum data')
    return HttpResponseRedirect(reverse("broker-zerodha-sum-list"))
