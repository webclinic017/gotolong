# Create your views here.

from .models import BrokerZerodhaTxn

from django.utils import timezone

from django.views.generic.list import ListView
from django.views.generic.dates import YearArchiveView, MonthArchiveView

from django.db.models import IntegerField, F, ExpressionWrapper, fields, Max, Min, Sum, Count
from django.db.models.functions import (ExtractYear, Round, ExtractMonth)
from django.db.models.expressions import RawSQL

import calendar
import pandas as pd
import csv, io
import openpyxl
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseRedirect

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update


class BrokerZerodhaTxnListView(ListView):
    model = BrokerZerodhaTxn

    # if pagination is desired
    # paginate_by = 300
    queryset = BrokerZerodhaTxn.objects.all()

    # month_list = BrokerZerodhaTxn.objects.dates('txn_date', 'month')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context


def txn_date_iso(self, txn_date, date_format):
    month_name_abbr_to_num_dict = {name: num for num, name in enumerate(calendar.month_abbr) if num}

    # zerodha - mm/dd/yyyy
    try:
        if date_format == 'mm/dd/yyyy':
            txn_date_arr = txn_date.split('/')
            txn_month = txn_date_arr[0].strip()
            txn_day = txn_date_arr[1].strip()
            txn_year = txn_date_arr[2].strip()
        elif date_format == 'yyyy-mm-dd':
            txn_date_arr = txn_date.split('-')
            txn_month = txn_date_arr[1].strip()
            txn_day = txn_date_arr[2].strip()
            txn_year = txn_date_arr[0].strip()
        elif date_format == 'dd-mmm-yy':
            txn_date_arr = txn_date.split('-')
            txn_day = txn_date_arr[0].strip()
            txn_month = txn_date_arr[1].strip()
            txn_year = txn_date_arr[2].strip()
        else:
            print('give right format')

        if txn_month.isdigit():
            # get rid of leading 0 in month number
            txn_month = str(int(txn_month))
        else:
            # month name to number
            txn_month = str(month_name_abbr_to_num_dict[txn_month])
        txn_date_iso = txn_year + "-" + txn_month + "-" + txn_day
        # ignore rest
    except ValueError:
        print('ValueError ', txn_date)
    except IndexError:
        print('IndexError ', txn_date)

    return txn_date_iso

# one parameter named request
def BrokerZerodhaTxnUpload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "invalid-request.html"
    data = BrokerZerodhaTxn.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        if sheet_name != 'Data':
            print("sheet name changed to", sheet_name)

        # ignore top two line : Average Market Capitalization of listed companies during the six months ended
        # remove top two line from dataframe
        df = df.iloc[2:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # change column name of data frame
        columns_list = ['bit_stock_symbol', 'bit_company_name', 'bit_isin_code', 'bit_action', 'bit_quantity',
                        'bit_txn_price', 'bit_brokerage', 'bit_txn_charges', 'bit_stamp_duty',
                        'bit_segment', 'bit_stt', 'bit_remarks', 'bit_txn_date',
                        'bit_exchange', 'bit_unused1']
        df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        df[["avg_mcap"]] = df[["avg_mcap"]].astype(int)

        # drop columns that are not required
        # skip_columns_list = ['bse_mcap', 'nse_mcap', 'mse_symbol', 'mse_mcap']
        # df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
        return HttpResponseRedirect(reverse("broker-zerodha-txn-list"))

    # delete existing records
    print('Deleted existing BrokerZerodhaTxn data')
    BrokerZerodhaTxn.objects.all().delete()

    # setup a stream which is when we loop through each line we are able to handle a data in a stream

    io_string = io.StringIO(data_set)
    next(io_string)
    unique_id = 0
    for column in csv.reader(io_string, delimiter=',', quotechar='"'):
        unique_id += 1
        txn_date = column[0].strip()
        stock_symbol = column[1].strip()

        print(unique_id, column)

        # convert mm/dd/yyyy to YYYY-mm-dd
        date_format = 'yyyy-mm-dd'
        txn_date = txn_date_iso(request, txn_date, date_format)

        _, created = BrokerZerodhaTxn.objects.update_or_create(
            bzt_id=unique_id,
            bzt_user_id=request.user.id,
            bzt_tdate=txn_date,
            bzt_tsymbol=stock_symbol,
            bzt_exchange=column[2],
            bzt_segment=column[3],
            bzt_trade_type=column[4],
            bzt_quantity=column[5],
            bzt_price=column[6],
            bzt_order_id=column[7],
            bzt_trade_id=column[8],
            bzt_order_exec_time=column[9]
        )
    # context = {}
    # render(request, template, context)
    #
    lastrefd_update("broker-zerodha-txn")
    #
    print('Completed loading new BrokerZerodhaTxn data')
    return HttpResponseRedirect(reverse("broker-zerodha-txn-list"))


def BrokerZerodhaTxnReset(request):
    # delete existing records
    print('Cleared existing BrokerZerodhaTxn data')
    BrokerZerodhaTxn.objects.all().filter(bzs_user_id=request.user.id).delete()
    return HttpResponseRedirect(reverse("broker-zerodha-txn-list"))
