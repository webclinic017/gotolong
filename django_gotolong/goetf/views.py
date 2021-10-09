# Create your views here.

from .models import Goetf

from django.views.generic.list import ListView

from django.db.models import (Count)

from django.urls import reverse
from django.http import HttpResponseRedirect
import urllib3
import csv
import io

import re

import openpyxl

import pandas as pd

from django_gotolong.lastrefd.models import Lastrefd, lastrefd_update


class GoetfListView(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_AUM(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().order_by('-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_Type(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().order_by('goetf_type', '-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_Benchmark(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().order_by('goetf_benchmark', '-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_IMF(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().filter(goetf_type='Index').order_by('-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_Gold(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().filter(goetf_benchmark__contains='Gold').order_by('-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfListView_Nifty(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']
    queryset = Goetf.objects.all().exclude(goetf_benchmark__contains='Gold').order_by('-goetf_aum')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        refresh_url = Goetf_url()
        context["refresh_url"] = refresh_url
        return context


class GoetfIndustryView(ListView):
    model = Goetf
    # if pagination is desired
    # paginate_by = 300
    # filter_backends = [filters.OrderingFilter,]
    # ordering_fields = ['sno', 'nse_symbol']

    queryset = Goetf.objects.all().values('goetf_benchmark').annotate(scheme_count=Count('goetf_benchmark')). \
        order_by('-scheme_count')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # get count of Industries
        benchmarks_count = len(Goetf.objects.all().values('goetf_benchmark'). \
                               annotate(benchmarks_count=Count('goetf_benchmark', distinct=True)))
        context['benchmarks_count'] = benchmarks_count
        return context


def Goetf_url():
    url = 'https://archives.nseindia.com/content/mfund/ind_nifty500list.csv'

    return url


# one parameter named request
def Goetf_fetch(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()
    debug_level = 1
    print('fetch not supported')
    return HttpResponseRedirect(reverse("goetf-list"))


# one parameter named request
def Goetf_upload(request):
    # for quick debugging
    #
    # import pdb; pdb.set_trace()
    #
    # breakpoint()

    debug_level = 1
    # declaring template
    template = "goetf/mfund_list.html"
    data = Goetf.objects.all()

    # GET request returns the value of the data with the specified key.
    if request.method == "GET":
        return render(request, template)

    req_file = request.FILES['file']

    # let's check if it is a csv file

    if req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx'):
        # get worksheet name
        # print('temporary file path:', req_file.temporary_file_path)
        print(req_file)

        if True:
            wb = openpyxl.load_workbook(req_file)
            print(wb.sheetnames)
            sheet_name = wb.sheetnames[0]
            print(sheet_name)
            ws = wb[sheet_name]
            df = pd.DataFrame(ws.values)
        else:
            xl = pd.ExcelFile(req_file)
            if debug_level > 0:
                print(xl.sheet_names)
            # single worksheet - Data
            sheet_name = xl.sheet_names[0]
            df = xl.parse(sheet_name)

        # can be 'Data'
        # can be 'Average MCap Jan Jun 2020'
        if sheet_name != 'fund-performance':
            print("sheet name changed to", sheet_name)

        # ignore top 6 line : Value Research, Fund Performance
        # remove top six line from dataframe
        df = df.iloc[6:]

        if debug_level > 0:
            print("old columns : ")
            print(df.columns)

        # change column name of data frame
        columns_list = ["scheme_name", "benchmark", "nav_date", "nav_regular", "nav_direct",
                        "return_1y_pct_regular", "return_1y_pct_direct", "return_1y_pct_benchmark",
                        "return_3y_pct_regular", "return_3y_pct_direct", "return_3y_pct_benchmark",
                        "return_5y_pct_regular", "return_5y_pct_direct", "return_5y_pct_benchmark",
                        "return_10y_pct_regular", "return_10y_pct_direct", "return_10y_pct_benchmark",
                        "return_since_launch_regular", "return_since_launch_direct", "return_since_launch_benchmark",
                        "daily_aum"]

        df.columns = columns_list

        if debug_level > 0:
            print("new columns : ")
            print(df.columns)

        # Keep only top 1000 entries
        # df = df.iloc[:1000]

        # round avg_mcap
        # df = df.round({'avg_mcap' : 1})
        # covert to numeric
        # df[["avg_mcap"]] = df[["avg_mcap"]].apply(pd.to_numeric)
        df[["daily_aum"]] = df[["daily_aum"]].astype(int)

        # drop columns that are not required
        skip_columns_list = ["nav_date", "nav_regular", "nav_direct",
                             "return_1y_pct_regular", "return_1y_pct_direct", "return_1y_pct_benchmark",
                             "return_3y_pct_regular", "return_3y_pct_direct", "return_3y_pct_benchmark",
                             "return_5y_pct_regular", "return_5y_pct_direct", "return_5y_pct_benchmark",
                             "return_10y_pct_regular", "return_10y_pct_direct", "return_10y_pct_benchmark",
                             "return_since_launch_regular", "return_since_launch_direct",
                             "return_since_launch_benchmark", ]

        df.drop(skip_columns_list, axis=1, inplace=True)

        data_set = df.to_csv(header=True, index=False)

    if req_file.name.endswith('.csv'):
        data_set = req_file.read().decode('UTF-8')

    if not (req_file.name.endswith('.csv') or req_file.name.endswith('.xls') or req_file.name.endswith('.xlsx')):
        messages.error(request, req_file.name + ' : THIS IS NOT A XLS/XLSX/CSV FILE.')
        return HttpResponseRedirect(reverse("goetf-list"))

    # delete existing records
    print('Deleted existing Goetf data')
    Goetf.objects.all().delete()

    # setup a stream which is when we loop through each line we are able to handle a data in a stream

    io_string = io.StringIO(data_set)
    # skip top three rows
    next(io_string)
    next(io_string)
    next(io_string)

    skip_records = 0
    for column in csv.reader(io_string, delimiter=',', quotechar='"'):
        goetf_scheme = column[0].strip()
        if re.search('Index', goetf_scheme):
            goetf_type = 'Index'
        elif re.search('ETF', goetf_scheme):
            goetf_type = 'ETF'
        elif re.search('Fund', goetf_scheme):
            goetf_type = 'Index'
        else:
            goetf_type = 'Unknown'
        goetf_benchmark = column[1].strip()
        goetf_aum = column[2].strip()

        _, created = Goetf.objects.update_or_create(
            goetf_scheme=goetf_scheme,
            goetf_type=goetf_type,
            goetf_benchmark=goetf_benchmark,
            goetf_aum=goetf_aum
        )

    lastrefd_update("goetf")

    print('Skipped records', skip_records)
    print('Completed loading new Goetf data')
    return HttpResponseRedirect(reverse("goetf-list"))
